import tempfile
import zipfile

import openpyxl
import pyexcel
import csv

from openpyxl.cell.read_only import EmptyCell

from app import logger


def convert_xls_to_xlsx(xls_file):
    xlsx_file_name = tempfile.mkstemp(suffix='.xlsx')[1]
    try:
        pyexcel.save_book_as(file_name=xls_file, dest_file_name=xlsx_file_name)
    except Exception as e:
        logger.info(
            'Error converting {} from .xls to .xlsx: {}, {}'.format(xls_file,
                                                                    type(e),
                                                                    str(e)))
        return None
    return xlsx_file_name


def convert_spreadsheet_to_csv(spreadsheet, parsed=True):
    if spreadsheet.endswith('.xls'):
        spreadsheet = convert_xls_to_xlsx(spreadsheet)
        if spreadsheet is None:
            return None

    try:
        workbook = openpyxl.load_workbook(open(spreadsheet, "rb"),
                                          read_only=True)
    except (KeyError, zipfile.BadZipfile) as e:
        logger.info('{} could not be opened as a spreadsheet: {}, {}'.format(
            spreadsheet, type(e), str(e)))
        return None

    return _convert_parsed(workbook) if parsed else _convert_blind(workbook)


def _convert_parsed(workbook):
    csv_file_name = tempfile.mkstemp()[1]

    rows = []
    column_names = None

    for sheet_name in list(workbook.sheetnames):
        sheet = workbook[sheet_name]

        sheet_column_names = {}
        for i, column in enumerate(
                list(sheet.iter_rows(min_row=1, max_row=1))[0]):
            sheet_column_names[i] = column.value.lower().strip()

        if column_names is None:
            column_names = list(sheet_column_names.values())
        elif set(sheet_column_names.values()).difference(set(column_names)):
            raise ValueError('all worksheets must contain the same columns')

        cell_types = []
        for row_cells in sheet.iter_rows(min_row=2):
            row = {}
            for column in range(0, len(sheet_column_names)):
                row[sheet_column_names[column]] = row_cells[
                                                      column].value or None
                cell_types.append(row_cells[column].data_type)
            rows.append(row)

        formula_cells = cell_types.count('f')
        if formula_cells > 0:
            raise RuntimeError(
                'Uploaded files can not contain formulas. Found {} cells with formulas.'.format(
                    formula_cells))

    with open(csv_file_name, 'w') as csv_file:
        writer = csv.DictWriter(csv_file, column_names)

        writer.writeheader()
        for row in rows:
            writer.writerow(row)

    return [csv_file_name]


def _truncate_row(row: tuple, n_empty=100):
    count = 0
    for i, cell in enumerate(row):
        if isinstance(cell, EmptyCell):
            count += 1
        else:
            count = 0

        if count == n_empty:
            return row[:i - n_empty - 1]
    return row


def _convert_blind(workbook):
    csv_file_names = []

    for sheet_name in list(workbook.sheetnames):
        sheet = workbook[sheet_name]
        csv_file_name = tempfile.mkstemp()[1]
        with open(csv_file_name, 'w', encoding='utf-8') as csv_file:
            writer = csv.writer(csv_file, delimiter=',')
            for i, row in enumerate(sheet.iter_rows(min_row=1)):
                if i % 100 == 0:
                    print(f'Processed {i + 1}/{sheet.max_row} rows')
                if 'f' in [w.data_type for w in row]:
                    raise RuntimeError(
                        'Uploaded files can not contain formulas')
                writer.writerow([cell.value for cell in _truncate_row(row)])

        csv_file_names.append(csv_file_name)

    return csv_file_names
