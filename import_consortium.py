# coding: utf-8

import openpyxl
import unicodecsv as csv
import argparse
from sqlalchemy.dialects.postgresql import JSONB
from sqlalchemy.sql import text
import datetime
import re
from dateutil import parser
from werkzeug.security import generate_password_hash

from app import db
from app import get_db_cursor
from account import Account
from account_grid_id import AccountGridId
from util import read_csv_file
from util import safe_commit
from util import is_issn



def import_grid_ids():
    rows = read_csv_file("/Users/hpiwowar/Downloads/SUNYgridids.csv")
    for row in rows:
        print row["username"], row["grid_id"]

        new_account = Account()
        new_account.username = row["username"]
        new_account.password_hash = generate_password_hash(row["username"])
        new_account.display_name = row["campus"]
        new_account.is_consortium = True
        new_account.consortium_id = "93YfzkaA"

        new_grid_id_object = AccountGridId()
        new_grid_id_object.grid_id = row["grid_id"]
        new_account.grid_id_objects.append(new_grid_id_object)

        # add AccountPackage things here
        # new_grid_id_object = AccountPackage()
        # insert into jump_account_package (select id, fn_short_uuid(), 'Elsevier', display_name || ' Elsevier', sysdate from jump_account where id not in (select package_id from jump_account_package))
        # new_grid_id_object.grid_id = row["grid_id"]
        # new_account.grid_id_objects.append(new_grid_id_object)

        # update jump_apc_authorships_view to use grid_id
        # jump_citing
        # jump_authorship
        # jump_counter

        db.session.add(new_account)

        safe_commit(db)


def import_consortium_counter(xls_filename):
    results = []
    xlsx_file = open(xls_filename, "rb")
    workbook = openpyxl.load_workbook(xlsx_file, read_only=True)
    sheetnames = list(workbook.sheetnames)

    for sheetname in sheetnames:
        university = sheetname.replace("JR1 ", "")
        print university

        sheet = workbook[sheetname]

        column_names = {}
        for i, column in enumerate(list(sheet.iter_rows(min_row=1, max_row=1))[0]):
            column_names[column.value] = i

        for row_cells in sheet.iter_rows(min_row=1):
            issn = row_cells[column_names['Print ISSN']].value
            total = row_cells[column_names['Reporting period total']].value
            if is_issn(issn):
                results.append({
                    "university": university,
                    "issn": issn,
                    "total": total
                })
                # print ".",
                # print university, issn, total

        with open("data/countercleaned.csv", "w") as csv_file:
            csv_writer = csv.writer(csv_file, encoding='utf-8')
            header = ["university", "issn", "total"]
            csv_writer.writerow(header)
            for my_dict in results:
                # doing this hacky thing so excel doesn't format the issn as a date :(
                csv_writer.writerow([my_dict[k] for k in header])


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Run stuff.")
    parser.add_argument('filename', type=str, help="xls file to parse")

    parsed_args = parser.parse_args()
    parsed_vars = vars(parsed_args)

    # import_grid_ids()
    import_consortium_counter(parsed_vars["filename"])




