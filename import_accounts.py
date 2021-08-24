# coding: utf-8

import openpyxl
import unicodecsv as csv
import argparse
from sqlalchemy.dialects.postgresql import JSONB
from sqlalchemy.sql import text
import datetime
import re
from dateutil import parser
from werkzeug.security import generate_password_hash, check_password_hash
import shortuuid
import glob
import codecs

from app import db
from app import get_db_cursor
from package import Package
from saved_scenario import SavedScenario
from util import read_csv_file
from util import safe_commit
from util import is_issn
from util import write_to_tempfile


def convert_from_utf16_to_utf8(filename):
    # https://stackoverflow.com/a/10300007/596939
    BLOCKSIZE = 1048576 # or some other, desired size in bytes
    with codecs.open(filename, "r", "utf-16") as sourceFile:
        new_filename = "{}_utf8".format(filename)
        with codecs.open(new_filename, "w", "utf-8") as targetFile:
            while True:
                contents = sourceFile.read(BLOCKSIZE)
                if not contents:
                    break
                targetFile.write(contents)
    return new_filename

def convert_date_spaces_to_dashes(filename):
    # https://stackoverflow.com/a/10300007/596939
    BLOCKSIZE = 1048576 # or some other, desired size in bytes
    with codecs.open(filename, "r", "utf-8") as sourceFile:
        new_filename = "{}_dashes".format(filename)
        with codecs.open(new_filename, "w", "utf-8") as targetFile:
            while True:
                contents = sourceFile.read(BLOCKSIZE)
                if not contents:
                    break
                contents = contents.replace(" 2018,", "-2018,")
                targetFile.write(contents)
    return new_filename



def import_consortium_counter_xls(xls_filename):
    results = []
    xlsx_file = open(xls_filename, "rb")
    workbook = openpyxl.load_workbook(xlsx_file, read_only=True)
    sheetnames = list(workbook.sheetnames)

    for sheetname in sheetnames:
        university = sheetname.replace("JR1 ", "")
        print(university)

        sheet = workbook[sheetname]

        column_names = {}
        for i, column in enumerate(list(sheet.iter_rows(min_row=1, max_row=1))[0]):
            column_names[column.value] = i

        for row_cells in sheet.iter_rows(min_row=1):
            issn = row_cells[column_names["Print ISSN"]].value
            total = row_cells[column_names["Reporting period total"]].value
            if is_issn(issn):
                results.append({
                    "university": university,
                    "issn": issn,
                    "total": total
                })
                # print ".",
                # print university, issn, total

        with open("data/countercleaned.csv", "w") as csv_file:
            csv_writer = csv.writer(csv_file, encoding="utf-8")
            header = ["university", "issn", "total"]
            csv_writer.writerow(header)
            for my_dict in results:
                # doing this hacky thing so excel doesn't format the issn as a date :(
                csv_writer.writerow([my_dict[k] for k in header])


def import_perpetual_access_files():
    results = []
    my_files = glob.glob("/Users/hpiwowar/Downloads/wvu_perpetual_access.csv")
    my_files.reverse()
    for my_file in my_files:
        print(my_file)
        if False:
            xlsx_file = open(my_file, "rb")
            workbook = openpyxl.load_workbook(xlsx_file, read_only=True)
            sheetnames = list(workbook.sheetnames)

            for sheetname in sheetnames:
                sheet = workbook[sheetname]

                column_names = {}
                for i, column in enumerate(list(sheet.iter_rows(min_row=1, max_row=1))[0]):
                    column_names[column.value] = i

                for row_cells in sheet.iter_rows(min_row=1):
                    username = row_cells[column_names["Account Name"]].value
                    issn = row_cells[column_names["ISSN (FS split)"]].value
                    start_date = row_cells[column_names["Content Start Date"]].value
                    end_date = row_cells[column_names["Content End Date"]].value
                    if is_issn(issn):
                        new_dict = {
                            "username": username,
                            "issn": issn,
                            "start_date": start_date,
                            "end_date": end_date
                        }
                        results.append(new_dict)
                        # print new_dict
                        print(".", end=' ')
        else:
            rows = read_csv_file(my_file)
            for row in rows:
                print(row)
                new_dict = {
                    "username": "wvu",
                    "issn": row["issn"],
                    "start_date": row["start_date"],
                    "end_date": row["end_date"]
                }
                results.append(new_dict)
                # print new_dict
                print(".", end=' ')

    with open("/Users/hpiwowar/Downloads/perpetual_access_cleaned.csv", "w") as csv_file:
        csv_writer = csv.writer(csv_file, encoding="utf-8")
        header = ["username", "issn", "start_date", "end_date"]
        csv_writer.writerow(header)
        for my_dict in results:
            csv_writer.writerow([my_dict[k] for k in header])
    print("/Users/hpiwowar/Downloads/perpetual_access_cleaned.csv")





# python import_accounts.py --filename=~/Downloads/new_accounts.csv
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Run stuff.")
    parser.add_argument("--filename", type=str, default=None, help="input file to parse")
    parser.add_argument("--username", type=str, default=None, help="username to input")

    parsed_args = parser.parse_args()
    parsed_vars = vars(parsed_args)

    # create_accounts(parsed_vars["filename"])
    # build_counter_import_file(filename=parsed_vars["filename"], username=parsed_vars["username"])


    crkn_ids = read_csv_file("/Users/hpiwowar/Documents/Projects/tiv2/jump-api/data/crkn_lookup.csv")
    institution_for_all_these_packages = "institution-fVnPvXK9iBYA"

    # report_name = "trj2"
    # all_in_one_data_rows = read_csv_file("/Users/hpiwowar/Documents/Projects/tiv2/jump-api/data/counter5_crkn/TR_J2 SUSHI Harvester CRKN_Wiley-2019_incl-non-participants.csv")

    # report_name = "trj3"
    # all_in_one_data_rows = read_csv_file("/Users/hpiwowar/Documents/Projects/tiv2/jump-api/data/counter5_crkn/TR_J3 SUSHI Harvester CRKN_Wiley-2019.csv")

    # report_name = "trj4"
    # all_in_one_data_rows = read_csv_file("/Users/hpiwowar/Documents/Projects/tiv2/jump-api/data/counter5_crkn/1 TR_J4 SUSHI Harvester CRKN_Wiley-2019.csv")
    # all_in_one_data_rows = read_csv_file("/Users/hpiwowar/Documents/Projects/tiv2/jump-api/data/counter5_crkn/2 TR_J4 SUSHI Harvester CRKN_Wiley-2019.csv")

    report_name = "trj4"
    all_in_one_data_rows = read_csv_file(" /Users/hpiwowar/Dropbox/companywide/unsub_customer_data/learn_counter5/raw_from_unis/crkn/elsevier/COP5_TR_J4_2019_UIR.xlsx")

    for row in crkn_ids:
        print("row: {}".format(list(row.values())))

        publisher_name = "Elseiver"
        crkn_institution_id = row["crkn_elsevier"]



        institution_number = row["institution_id"].replace("institution-", "")
        package_id = "package-CRKNcounter5{}{}".format(publisher_name, institution_number)
        package_name = "CRKN COUNTER5 {} {}".format(publisher_name, row["name"])

        from counter import CounterInput
        my_counter_test = CounterInput.query.filter(CounterInput.package_id == package_id,
                                                    CounterInput.report_name == report_name).first()
        has_counter_data_to_load = False

        if my_counter_test:
            print("counter already loaded for {}".format(report_name))
        else:
            temp_filename = "data/temp.csv"
            with open(temp_filename, "w") as csv_file:
                csv_writer = csv.writer(csv_file, encoding="utf-8")
                header = list(all_in_one_data_rows[0].keys())
                csv_writer.writerow(header)
                for row in all_in_one_data_rows:
                    if row["Customer ID"] == crkn_institution_id:
                        # doing this hacky thing so excel doesn't format the issn as a date :(
                        csv_writer.writerow([row[k] for k in header])
                        has_counter_data_to_load = True

            print("now loading in counter data")
            if not has_counter_data_to_load:
                print("has no counter data to load")
            else:
                # see if package exists:
                my_package = Package.query.get(package_id)
                if not my_package:
                    my_package = Package(
                        package_id=package_id,
                        publisher=publisher_name,
                        package_name=package_name,
                        created=datetime.datetime.utcnow().isoformat(),
                        institution_id=institution_for_all_these_packages,
                        is_demo=False
                    )
                    db.session.add(my_package)
                    safe_commit(db)
                else:
                    print("package already created")

                print("{}".format(my_package))

                CounterInput().load(package_id, temp_filename, commit=True)





